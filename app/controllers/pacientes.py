import io
import re
import zipfile
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import PurePosixPath
from threading import Lock
from uuid import uuid4

import openpyxl
from flask import Blueprint, flash, jsonify, redirect, render_template, request, url_for
from flask_login import current_user
from sqlalchemy import func, or_
from sqlalchemy.exc import IntegrityError

from app import db_orm as db
from app.core.audit import AuditLog
from app.core.auth import login_required
from app.core.security import roles_required
from app.core.validators import (
    ALLOWED_APPOINTMENT_STATUS,
    ValidationError,
    appointment_payload,
    assessment_payload,
    clean_text,
    date_value,
    enum_value,
    integer,
    number,
    patient_payload,
    payment_payload,
)
from app.models.cita import Cita
from app.models.historial_clinico import HistorialClinico
from app.models.paciente import Paciente
from app.models.pago import Pago
from app.models.valoracion_antropometrica import ValoracionAntropometrica

pacientes = Blueprint("pacientes", __name__, url_prefix="/pacientes")
agenda = Blueprint("agenda", __name__, url_prefix="/agenda")
MAX_EXCEL_ROWS = 2000
MAX_XLSX_UNCOMPRESSED = 25 * 1024 * 1024
APPOINTMENT_CALENDAR_DAYS = 21
APPOINTMENT_MAX_FUTURE_DAYS = 730
WEEKDAYS_SHORT_ES = ("Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom")
WEEKDAYS_ES = ("Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo")
MONTHS_SHORT_ES = ("Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic")
MONTHS_ES = (
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
)
_APPOINTMENT_WRITE_LOCK = Lock()


def _flash_validation(error):
    flash(str(error), "error")


def _appointment_calendar(moment=None, exclude_appointment_id=None):
    current = moment or datetime.now()
    start = current.date()
    end = start + timedelta(days=APPOINTMENT_CALENDAR_DAYS - 1)
    occupied = {day: set() for day in (start + timedelta(days=offset) for offset in range(APPOINTMENT_CALENDAR_DAYS))}
    query = Cita.query.with_entities(Cita.fecha, Cita.hora).filter(
        Cita.estatus == "Programada", Cita.fecha.between(start, end)
    )
    if exclude_appointment_id:
        query = query.filter(Cita.id != exclude_appointment_id)
    rows = query.all()
    for appointment_date, appointment_time in rows:
        occupied.setdefault(appointment_date, set()).add(appointment_time)

    days = []
    for offset in range(APPOINTMENT_CALENDAR_DAYS):
        target = start + timedelta(days=offset)
        available = sum(
            1
            for slot in Cita.HORARIOS_ATENCION
            if slot not in occupied.get(target, set()) and datetime.combine(target, slot) > current
        )
        days.append(
            {
                "fecha": target,
                "iso": target.isoformat(),
                "dia_semana": WEEKDAYS_SHORT_ES[target.weekday()],
                "mes": MONTHS_SHORT_ES[target.month - 1],
                "disponibles": available,
            }
        )
    return days


def _quick_appointment_context(appointment=None):
    editing = appointment is not None
    calendar = _appointment_calendar(exclude_appointment_id=appointment.id if editing else None)
    raw_date = (
        request.form.get("proxima_cita_fecha")
        or request.args.get("fecha")
        or (appointment.fecha.isoformat() if editing else None)
    )
    selected_date = None
    if raw_date:
        try:
            candidate = date_value(raw_date, "Fecha de cita", allow_future=True)
            if date.today() <= candidate <= date.today() + timedelta(days=APPOINTMENT_MAX_FUTURE_DAYS):
                selected_date = candidate
        except ValidationError:
            selected_date = None
    if not selected_date:
        first_available = next((item for item in calendar if item["disponibles"]), calendar[0])
        selected_date = first_available["fecha"]
    selected_patient = appointment.paciente if editing else None
    pending_appointment = None
    raw_patient_id = (
        appointment.paciente_id
        if editing
        else request.form.get("paciente_id") or request.args.get("paciente_id", "")
    )
    if raw_patient_id:
        try:
            patient_id = integer(raw_patient_id, "Paciente", minimum=1)
            candidate = db.session.get(Paciente, patient_id)
            if candidate and candidate.status == "activo":
                selected_patient = candidate
                pending_appointment = None if editing else Cita.obtener_cita_pendiente(candidate.id)
        except ValidationError:
            pass
    return {
        "calendario_citas": calendar,
        "fecha_seleccionada": selected_date,
        "paciente_seleccionado": selected_patient,
        "cita_pendiente_seleccionada": pending_appointment,
        "fecha_minima": date.today(),
        "fecha_maxima": date.today() + timedelta(days=APPOINTMENT_MAX_FUTURE_DAYS),
        "cita_edicion": appointment,
        "modo_reagenda": editing,
        "origen_agenda": editing
        or request.form.get("origen") == "agenda"
        or request.args.get("origen") == "agenda",
    }


def _appointment_patient_result(patient, pending_appointment=None):
    appointment_data = None
    if pending_appointment:
        appointment_data = {
            "fecha": pending_appointment.fecha.isoformat(),
            "hora": pending_appointment.hora.strftime("%H:%M"),
            "etiqueta": (
                f"{pending_appointment.fecha.strftime('%d/%m/%Y')} · "
                f"{pending_appointment.hora.strftime('%H:%M')}"
            ),
        }
    return {
        "id": patient.id,
        "nombre": patient.nombre_completo,
        "expediente": f"EXP-{patient.id:04d}",
        "telefono": patient.telefono,
        "detalle_url": url_for("pacientes.detalle_paciente", id=patient.id),
        "cita_programada": appointment_data,
    }


@pacientes.route("/nuevo", methods=["GET", "POST"])
@login_required
def nuevo_paciente():
    if request.method == "POST":
        try:
            data = patient_payload(request.form)
            patient = Paciente.crear(**data, status="activo")
            db.session.flush()
            AuditLog.record("paciente.create", entity_type="paciente", entity_id=patient.id)
            db.session.commit()
            flash("Paciente registrado exitosamente.", "success")
            return redirect(url_for("pacientes.lista_pacientes_activos"))
        except ValidationError as error:
            _flash_validation(error)
        except IntegrityError:
            db.session.rollback()
            flash("No fue posible registrar al paciente por un dato duplicado.", "error")
    return render_template("pacientes/nuevo_paciente.html")


def _patient_list(status):
    search = str(request.args.get("busqueda", ""))[:100]
    order_by = request.args.get("ordenar_por", "id")
    order = request.args.get("orden", "desc")
    if order_by not in {"id", "nombre", "apellidos", "ultima_consulta"}:
        order_by = "id"
    if order not in {"asc", "desc"}:
        order = "desc"
    patients = Paciente.buscar(search, status=status, ordenar_por=order_by, orden=order)
    return render_template(
        "pacientes/lista_pacientes.html",
        pacientes=patients,
        busqueda=search,
        ordenar_por=order_by,
        orden=order,
        tipo_lista="activos" if status == "activo" else "inactivos",
    )


@pacientes.route("/activos")
@login_required
def lista_pacientes_activos():
    return _patient_list("activo")


@pacientes.route("/inactivos")
@login_required
def lista_pacientes_inactivos():
    return _patient_list("inactivo")


@pacientes.route("/<int:id>")
@login_required
def detalle_paciente(id):
    patient = db.session.get(Paciente, id)
    if not patient:
        flash("Paciente no encontrado.", "error")
        return redirect(url_for("pacientes.lista_pacientes_activos"))
    can_view_clinical = current_user.rol_clinico in {"admin", "medico"}
    assessments = ValoracionAntropometrica.obtener_por_paciente(id) if can_view_clinical else []
    current = assessments[0] if assessments else None
    previous = assessments[1] if len(assessments) > 1 else None
    differences = {}
    if current and previous:
        for field in [
            "cintura",
            "torax",
            "brazo",
            "bicep",
            "tricep",
            "cadera",
            "pierna",
            "pantorrilla",
            "subescapular",
            "suprailiaco",
            "femoral",
        ]:
            left, right = getattr(current, field), getattr(previous, field)
            if left is not None and right is not None:
                delta = left - right
                differences[field] = {
                    "valor": delta,
                    "tendencia": "aumento" if delta > 0 else "reduccion" if delta < 0 else "sin_cambio",
                }
    next_appointment = Cita.obtener_siguiente_cita(id)
    next_datetime = datetime.combine(next_appointment.fecha, next_appointment.hora) if next_appointment else None
    payment_appointments = (
        Cita.query.filter_by(paciente_id=id)
        .order_by(Cita.fecha.desc(), Cita.hora.desc(), Cita.id.desc())
        .limit(20)
        .all()
    )
    return render_template(
        "pacientes/detalle_paciente.html",
        paciente=patient,
        ultimo_pago=Pago.obtener_ultimo_pago(id),
        pagos_historial=Pago.obtener_historial_paciente(id),
        payment_operation_key=str(uuid4()),
        fecha_pago_default=date.today(),
        fecha_pago_desde_default=date.today() - timedelta(days=365),
        citas_pago=payment_appointments,
        historial=HistorialClinico.obtener_por_paciente_id(id) if can_view_clinical else None,
        ultima_valoracion=current,
        valoracion_anterior=previous,
        diferencias=differences,
        siguiente_cita=next_datetime,
        cita_programada=next_appointment,
        today=datetime.now(),
    )


@pacientes.route("/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_paciente(id):
    patient = db.session.get(Paciente, id)
    if not patient:
        flash("Paciente no encontrado.", "error")
        return redirect(url_for("pacientes.lista_pacientes_activos"))
    if request.method == "POST":
        try:
            data = patient_payload(request.form, include_status=True)
            for key, value in data.items():
                setattr(patient, key, value)
            AuditLog.record(
                "paciente.update", entity_type="paciente", entity_id=patient.id, metadata={"status": patient.status}
            )
            db.session.commit()
            flash("Paciente actualizado exitosamente.", "success")
            return redirect(url_for("pacientes.detalle_paciente", id=id))
        except ValidationError as error:
            _flash_validation(error)
    return render_template("pacientes/editar_paciente.html", paciente=patient)


@pacientes.route("/<int:id>/pago", methods=["POST"])
@login_required
def registrar_pago_paciente(id):
    patient = db.session.get(Paciente, id)
    if not patient:
        flash("Paciente no encontrado.", "error")
        return redirect(url_for("pacientes.lista_pacientes_activos"))
    try:
        data = payment_payload(request.form)
        appointment_id = None
        raw_appointment_id = str(request.form.get("cita_id") or "").strip()
        if raw_appointment_id:
            appointment_id = integer(raw_appointment_id, "Cita", minimum=1)
            appointment = db.session.get(Cita, appointment_id)
            if not appointment or appointment.paciente_id != patient.id:
                raise ValidationError("La cita seleccionada no pertenece al paciente.")
        payment = Pago.crear(
            id,
            data,
            usuario_id=current_user.id,
            cita_id=appointment_id,
        )
        db.session.flush()
        AuditLog.record(
            "pago.create",
            entity_type="pago",
            entity_id=payment.id,
            metadata={"paciente_id": id, "folio": payment.folio, "cita_id": appointment_id},
        )
        db.session.commit()
        flash(f"Pago {payment.folio} registrado exitosamente.", "success")
    except ValidationError as error:
        db.session.rollback()
        _flash_validation(error)
    except IntegrityError:
        db.session.rollback()
        duplicate = Pago.query.filter_by(operation_key=str(request.form.get("operation_key") or "")).first()
        AuditLog.record(
            "pago.duplicate",
            entity_type="pago",
            entity_id=duplicate.id if duplicate else None,
            outcome="denied",
            metadata={"paciente_id": id},
        )
        db.session.commit()
        if duplicate and duplicate.paciente_id == patient.id:
            flash(f"Este pago ya se registró con el folio {duplicate.folio}. No se creó otro.", "warning")
        else:
            flash("No fue posible registrar el pago. Recarga el formulario e inténtalo nuevamente.", "error")
    return redirect(url_for("pacientes.detalle_paciente", id=id))


@pacientes.route("/<int:id>/cambiar-estado", methods=["POST"])
@login_required
def cambiar_estado(id):
    patient = db.session.get(Paciente, id)
    if not patient:
        return jsonify({"success": False, "error": "Paciente no encontrado"}), 404
    patient.status = "inactivo" if patient.status == "activo" else "activo"
    AuditLog.record(
        "paciente.status", entity_type="paciente", entity_id=patient.id, metadata={"status": patient.status}
    )
    db.session.commit()
    return jsonify({"success": True, "nuevo_estado": patient.status})


def _excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except (TypeError, ValueError):
            continue
    raise ValidationError("Revisa la fecha; usa una fecha válida.")


def _extract_number(text, key):
    aliases = {
        "tc": "(?:tc|t)",
        "bc": "(?:bc|b)",
        "si": "(?:si|i)",
        "se": "(?:se|e)",
        "fem": "(?:fem|f)",
        "grasa": "grasa",
    }
    match = re.search(rf"{aliases[key]}\s*[:=]?\s*(\d+(?:[.,]\d+)?)", str(text or "").lower())
    return match.group(1) if match else None


def _validate_xlsx_archive(blob):
    stream = io.BytesIO(blob)
    if not zipfile.is_zipfile(stream):
        raise ValidationError("El archivo no se puede leer como Excel. Guarda una copia en formato .xlsx e inténtalo de nuevo.")
    stream.seek(0)
    with zipfile.ZipFile(stream) as archive:
        members = archive.infolist()
        if len(members) > 500:
            raise ValidationError("El archivo de Excel no se puede procesar. Revisa que no esté dañado y vuelve a guardarlo como .xlsx.")
        total = 0
        for member in members:
            path = PurePosixPath(member.filename)
            if path.is_absolute() or ".." in path.parts:
                raise ValidationError("El archivo de Excel no se puede procesar. Revisa que no esté dañado y vuelve a guardarlo como .xlsx.")
            total += member.file_size
            if member.file_size > 10 * 1024 * 1024:
                raise ValidationError("El archivo de Excel no se puede procesar. Revisa que no esté dañado y vuelve a guardarlo como .xlsx.")
            if member.file_size and member.compress_size == 0:
                raise ValidationError("El archivo de Excel no se puede procesar. Revisa que no esté dañado y vuelve a guardarlo como .xlsx.")
            if member.compress_size and member.file_size / member.compress_size > 100:
                raise ValidationError("El archivo de Excel no se puede procesar. Revisa que no esté dañado y vuelve a guardarlo como .xlsx.")
        if total > MAX_XLSX_UNCOMPRESSED:
            raise ValidationError("El archivo de Excel no se puede procesar. Revisa que no esté dañado y vuelve a guardarlo como .xlsx.")


@pacientes.route("/<int:id>/cargar-excel", methods=["POST"])
@login_required
@roles_required("admin", "medico")
def cargar_excel(id):
    if not current_user.puede_capturar_antropometria:
        AuditLog.record(
            "valoracion.import",
            entity_type="paciente",
            entity_id=id,
            outcome="denied",
            metadata={
                "reason": "professional_profile_not_allowed",
                "perfil_profesional": current_user.perfil_profesional_clinico,
            },
        )
        db.session.commit()
        return jsonify(
            {
                "success": False,
                "message": "La importación antropométrica sólo está disponible para profesionales de Nutrición.",
            }
        ), 403
    if not db.session.get(Paciente, id):
        return jsonify({"success": False, "message": "Paciente no encontrado"}), 404
    uploaded = request.files.get("excel_file") or request.files.get("file")
    if not uploaded or not uploaded.filename:
        return jsonify({"success": False, "message": "Selecciona un archivo de Excel (.xlsx)."}), 400
    if not uploaded.filename.lower().endswith(".xlsx"):
        return jsonify({"success": False, "message": "Elige un archivo de Excel guardado en formato .xlsx."}), 400
    blob = uploaded.read(5 * 1024 * 1024 + 1)
    if len(blob) > 5 * 1024 * 1024:
        return jsonify({"success": False, "message": "El archivo es demasiado grande. Elige uno de hasta 5 MB."}), 413
    try:
        _validate_xlsx_archive(blob)
        workbook = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True, keep_links=False)
        sheet = workbook.active
        if sheet.max_row > MAX_EXCEL_ROWS or sheet.max_column > 100:
            raise ValidationError(f"El archivo tiene demasiadas filas o columnas. Usa hasta {MAX_EXCEL_ROWS} filas y vuelve a intentarlo.")
        height_raw = sheet["M8"].value
        height = number(height_raw, "Estatura", minimum=0.5, maximum=250)
        if height > 3:
            height = height / 100

        existing = {
            (row.numero_cita, row.fecha) for row in ValoracionAntropometrica.query.filter_by(paciente_id=id).all()
        }
        seen, pending, duplicates, errors = set(), [], 0, []
        for row_number in range(10, sheet.max_row + 1):
            visit_raw = sheet.cell(row=row_number, column=12).value
            date_raw = sheet.cell(row=row_number, column=13).value
            weight_raw = sheet.cell(row=row_number, column=14).value
            if visit_raw in (None, "") and date_raw in (None, "") and weight_raw in (None, ""):
                continue
            try:
                visit_number = number(visit_raw, "Número de cita", minimum=1, maximum=10000)
                if not visit_number.is_integer():
                    raise ValidationError("Revisa la columna Número de cita; debe contener números sin decimales.")
                visit_number = int(visit_number)
                assessment_date = _excel_date(date_raw)
                key = (visit_number, assessment_date)
                if key in existing or key in seen:
                    duplicates += 1
                    continue
                seen.add(key)
                composite = sheet.cell(row=row_number, column=15).value
                folds = sheet.cell(row=row_number, column=22).value
                percent_raw = str(sheet.cell(row=row_number, column=23).value or "").replace("%", "")
                raw = {
                    "numero_cita": visit_number,
                    "fecha": assessment_date.strftime("%Y-%m-%d"),
                    "motivo_consulta": "Consulta importada desde expediente antropométrico",
                    "estatura": height,
                    "peso": weight_raw,
                    "grasa": _extract_number(composite, "grasa") or 0,
                    "cintura": sheet.cell(row=row_number, column=16).value,
                    "torax": sheet.cell(row=row_number, column=17).value,
                    "brazo": sheet.cell(row=row_number, column=18).value,
                    "cadera": sheet.cell(row=row_number, column=19).value,
                    "pierna": sheet.cell(row=row_number, column=20).value,
                    "pantorrilla": sheet.cell(row=row_number, column=21).value,
                    "tension_arterial": sheet.cell(row=row_number, column=24).value,
                    "frecuencia_cardiaca": sheet.cell(row=row_number, column=25).value,
                    "bicep": _extract_number(folds, "bc"),
                    "tricep": _extract_number(folds, "tc"),
                    "suprailiaco": _extract_number(folds, "si"),
                    "subescapular": _extract_number(folds, "se"),
                    "femoral": _extract_number(folds, "fem"),
                    "porcentaje_grasa": percent_raw,
                }
                pending.append(assessment_payload(raw, allow_anthropometry=True))
            except ValidationError as error:
                errors.append(f"Fila {row_number}: {error}")
        workbook.close()
        if errors:
            return jsonify(
                {
                    "success": False,
                    "message": "Hay datos que necesitan corregirse. No se agregó ninguna consulta.",
                    "errores": errors[:20],
                }
            ), 400
        with ValoracionAntropometrica.bloqueo_numeracion_diaria():
            next_by_date = {}
            for data in pending:
                assessment_date = data["fecha"]
                if assessment_date not in next_by_date:
                    next_by_date[assessment_date] = ValoracionAntropometrica.siguiente_numero_diario(
                        assessment_date
                    )
                data["numero_cita"] = next_by_date[assessment_date]
                next_by_date[assessment_date] += 1
                ValoracionAntropometrica.crear(id, data, profesional=current_user)
            AuditLog.record(
                "valoracion.import",
                entity_type="paciente",
                entity_id=id,
                metadata={
                    "created": len(pending),
                    "duplicates": duplicates,
                    "daily_sequence": True,
                },
            )
            db.session.commit()
        return jsonify(
            {
                "success": True,
                "message": f"Se agregaron {len(pending)} consultas. {duplicates} ya existían y no se volvieron a agregar.",
                "registros_duplicados": duplicates,
                "registros_procesados": len(pending),
                "errores": [],
            }
        )
    except (ValidationError, zipfile.BadZipFile, KeyError) as error:
        db.session.rollback()
        return jsonify({"success": False, "message": str(error)}), 400
    except IntegrityError:
        db.session.rollback()
        return jsonify(
            {
                "success": False,
                "message": "No se pudieron ordenar las consultas importadas. Inténtalo de nuevo.",
            }
        ), 409
    except (OSError, ValueError, TypeError):
        db.session.rollback()
        return jsonify({"success": False, "message": "No se pudo leer el archivo de Excel. Revisa que sea un .xlsx válido."}), 400


def _appointment_values(form):
    data = appointment_payload(form)
    moment = datetime.combine(data["fecha"], data["hora"])
    if moment <= datetime.now():
        raise ValidationError("La cita debe programarse en una fecha y hora futuras.")
    return data


def _agenda_target_date():
    raw = request.args.get("fecha")
    if not raw:
        return date.today()
    target = date_value(raw, "Fecha de agenda", allow_future=True)
    if target > date.today() + timedelta(days=APPOINTMENT_MAX_FUTURE_DAYS):
        raise ValidationError("La agenda sólo puede consultarse hasta dos años en el futuro.")
    return target


def _agenda_context(target, view):
    view = view if view in {"dia", "semana"} else "dia"
    start = target if view == "dia" else target - timedelta(days=target.weekday())
    end = start if view == "dia" else start + timedelta(days=6)
    appointments = Cita.obtener_periodo(start, end)
    by_day = defaultdict(list)
    now = datetime.now()
    for appointment in appointments:
        appointment_moment = datetime.combine(appointment.fecha, appointment.hora)
        active_patient = bool(appointment.paciente and appointment.paciente.status == "activo")
        appointment.puede_reagendar = (
            appointment.estatus == "Programada" and appointment_moment > now and active_patient
        )
        appointment.puede_cerrar = appointment.estatus == "Programada" and appointment_moment <= now
        appointment.puede_iniciar_consulta = (
            appointment.estatus == "Programada" and appointment.fecha == now.date() and active_patient
        )
        by_day[appointment.fecha].append(appointment)

    days = []
    for offset in range((end - start).days + 1):
        day = start + timedelta(days=offset)
        rows = by_day.get(day, [])
        days.append(
            {
                "fecha": day,
                "dia_semana": WEEKDAYS_ES[day.weekday()],
                "mes": MONTHS_ES[day.month - 1],
                "citas": rows,
                "programadas": sum(1 for item in rows if item.estatus == "Programada"),
            }
        )

    step = 1 if view == "dia" else 7
    counts = {status: sum(1 for item in appointments if item.estatus == status) for status in ALLOWED_APPOINTMENT_STATUS}
    range_label = (
        f"{WEEKDAYS_ES[start.weekday()]} {start.day} de {MONTHS_ES[start.month - 1]} de {start.year}"
        if view == "dia"
        else (
            f"{start.day} de {MONTHS_ES[start.month - 1]}"
            f" — {end.day} de {MONTHS_ES[end.month - 1]} de {end.year}"
        )
    )
    return {
        "vista_agenda": view,
        "fecha_objetivo": target,
        "fecha_hoy": date.today(),
        "fecha_inicio": start,
        "fecha_fin": end,
        "fecha_anterior": start - timedelta(days=step),
        "fecha_siguiente": start + timedelta(days=step),
        "fecha_maxima": date.today() + timedelta(days=APPOINTMENT_MAX_FUTURE_DAYS),
        "fecha_nueva_cita": max(target, date.today()),
        "rango_etiqueta": range_label,
        "dias_agenda": days,
        "total_citas": len(appointments),
        "conteos_estatus": counts,
        "can_view_clinical": current_user.rol_clinico in {"admin", "medico"},
    }


def _validate_appointment_status_transition(appointment, status, reason):
    if appointment.estatus != "Programada":
        raise ValidationError("La cita ya está cerrada y no admite otro cambio de estado.")
    if status == "Programada":
        raise ValidationError("La cita ya se encuentra programada.")
    if status in {"Atendida", "No Asistió"}:
        appointment_moment = datetime.combine(appointment.fecha, appointment.hora)
        if appointment_moment > datetime.now():
            raise ValidationError("Una cita futura no puede cerrarse como atendida o no asistida.")
    if status == "Cancelada" and not reason:
        raise ValidationError("Indica el motivo de cancelación.")


@agenda.route("")
@login_required
def index():
    try:
        target = _agenda_target_date()
    except ValidationError as error:
        flash(str(error), "warning")
        target = date.today()
    view = request.args.get("vista", "dia")
    if view not in {"dia", "semana"}:
        flash("La vista solicitada no es válida; se mostró la agenda del día.", "warning")
        view = "dia"
    return render_template("agenda/index.html", **_agenda_context(target, view))


@agenda.route("/citas/<int:cita_id>/reagendar", methods=["GET", "POST"])
@login_required
def reagendar_cita(cita_id):
    appointment = db.session.get(Cita, cita_id)
    if not appointment:
        flash("Cita no encontrada.", "error")
        return redirect(url_for("agenda.index"))
    if appointment.estatus != "Programada":
        flash("Sólo las citas programadas pueden reagendarse.", "warning")
        return redirect(url_for("agenda.index", fecha=appointment.fecha.isoformat()))
    if datetime.combine(appointment.fecha, appointment.hora) <= datetime.now():
        flash("Una cita cuyo horario ya transcurrió debe cerrarse como atendida o no asistida.", "warning")
        return redirect(url_for("agenda.index", fecha=appointment.fecha.isoformat()))
    if not appointment.paciente or appointment.paciente.status != "activo":
        flash("No es posible reagendar porque el paciente no está activo.", "warning")
        return redirect(url_for("agenda.index", fecha=appointment.fecha.isoformat()))

    if request.method == "POST":
        try:
            data = _appointment_values(request.form)
            if data["fecha"] > date.today() + timedelta(days=APPOINTMENT_MAX_FUTURE_DAYS):
                raise ValidationError("La cita no puede programarse con más de dos años de anticipación.")
            with _APPOINTMENT_WRITE_LOCK:
                if appointment.estatus != "Programada":
                    raise ValidationError("La cita cambió de estado y ya no puede reagendarse.")
                if not Cita.es_horario_disponible(data["fecha"], data["hora"], appointment.id):
                    raise ValidationError("El horario seleccionado ya no está disponible.")
                previous = {
                    "fecha": appointment.fecha.isoformat(),
                    "hora": appointment.hora.strftime("%H:%M"),
                }
                appointment.fecha = data["fecha"]
                appointment.hora = data["hora"]
                appointment.motivo = data["motivo"] or None
                appointment.estado = "pendiente"
                appointment.motivo_cancelacion = None
                AuditLog.record(
                    "cita.update",
                    entity_type="cita",
                    entity_id=appointment.id,
                    metadata={
                        "paciente_id": appointment.paciente_id,
                        "origen": "agenda",
                        "anterior": previous,
                        "fecha": appointment.fecha.isoformat(),
                        "hora": appointment.hora.strftime("%H:%M"),
                    },
                )
                db.session.commit()
            flash("Cita reagendada correctamente.", "success")
            return redirect(url_for("agenda.index", fecha=appointment.fecha.isoformat(), vista="dia"))
        except ValidationError as error:
            db.session.rollback()
            AuditLog.record(
                "cita.update",
                entity_type="cita",
                entity_id=appointment.id,
                outcome="denied",
                metadata={
                    "paciente_id": appointment.paciente_id,
                    "origen": "agenda",
                    "validacion": str(error)[:160],
                },
            )
            db.session.commit()
            _flash_validation(error)
            return render_template(
                "pacientes/agendar_cita.html", **_quick_appointment_context(appointment)
            ), 400
    return render_template("pacientes/agendar_cita.html", **_quick_appointment_context(appointment))


@pacientes.route("/buscar_para_cita", methods=["GET"])
@login_required
def buscar_para_cita():
    try:
        search = clean_text(request.args.get("busqueda"), "Búsqueda", maximum=100)
    except ValidationError as error:
        return jsonify({"success": False, "error": str(error)}), 400

    if len(search) < 2:
        response = jsonify({"success": True, "resultados": []})
        response.headers["Cache-Control"] = "no-store"
        return response

    from app.core.text import search_key, search_terms

    normalized = search_key(search)
    searchable_fields = (
        func.sgpn_search_key(Paciente.nombre),
        func.sgpn_search_key(Paciente.apellido_paterno),
        func.sgpn_search_key(Paciente.apellido_materno),
        func.sgpn_search_key(Paciente.telefono),
        func.sgpn_search_key(Paciente.correo),
    )
    terms = search_terms(search)
    filters = [
        or_(*(field.contains(term, autoescape=True) for field in searchable_fields))
        for term in terms
    ]
    record_match = re.fullmatch(r"exp[\s-]*0*(\d{1,9})", normalized)
    if record_match:
        filters = [Paciente.id == int(record_match.group(1))]
    elif not terms:
        response = jsonify({"success": True, "resultados": []})
        response.headers["Cache-Control"] = "no-store"
        return response

    query = Paciente.query.filter(Paciente.status == "activo")
    for term_filter in filters:
        query = query.filter(term_filter)
    matches = (
        query
        .order_by(Paciente.nombre.asc(), Paciente.apellido_paterno.asc(), Paciente.id.asc())
        .limit(8)
        .all()
    )
    pending_by_patient = {}
    if matches:
        pending = (
            Cita.query.filter(Cita.paciente_id.in_([patient.id for patient in matches]), Cita.estatus == "Programada")
            .order_by(Cita.fecha.asc(), Cita.hora.asc(), Cita.id.asc())
            .all()
        )
        for appointment in pending:
            pending_by_patient.setdefault(appointment.paciente_id, appointment)

    response = jsonify(
        {
            "success": True,
            "resultados": [
                _appointment_patient_result(patient, pending_by_patient.get(patient.id)) for patient in matches
            ],
        }
    )
    response.headers["Cache-Control"] = "no-store"
    return response


@pacientes.route("/agendar-cita", methods=["GET", "POST"])
@login_required
def agendar_cita_rapida():
    if request.method == "POST":
        try:
            origin = "agenda" if request.form.get("origen") == "agenda" else "kpi_dashboard"
            patient_id = integer(request.form.get("paciente_id"), "Paciente", minimum=1)
            patient = db.session.get(Paciente, patient_id)
            if not patient or patient.status != "activo":
                raise ValidationError("Selecciona un paciente activo registrado.")
            data = _appointment_values(request.form)
            if data["fecha"] > date.today() + timedelta(days=APPOINTMENT_MAX_FUTURE_DAYS):
                raise ValidationError("La cita no puede programarse con más de dos años de anticipación.")

            with _APPOINTMENT_WRITE_LOCK:
                pending = Cita.obtener_cita_pendiente(patient.id)
                if pending:
                    raise ValidationError(
                        "El paciente ya tiene una cita programada. Modifícala desde el detalle del paciente."
                    )
                if not Cita.es_horario_disponible(data["fecha"], data["hora"]):
                    raise ValidationError("El horario seleccionado ya no está disponible.")
                appointment = Cita(
                    paciente_id=patient.id,
                    fecha=data["fecha"],
                    hora=data["hora"],
                    motivo=data["motivo"] or None,
                    estado="pendiente",
                    estatus="Programada",
                )
                db.session.add(appointment)
                db.session.flush()
                AuditLog.record(
                    "cita.create",
                    entity_type="cita",
                    entity_id=appointment.id,
                    metadata={"paciente_id": patient.id, "origen": origin},
                )
                db.session.commit()
            if origin == "agenda":
                flash("Cita agendada correctamente.", "success")
                return redirect(url_for("agenda.index", fecha=appointment.fecha.isoformat(), vista="dia"))
            flash("Cita agendada exitosamente desde el Dashboard.", "success")
            return redirect(f"{url_for('main.index')}#agenda-hoy")
        except ValidationError as error:
            _flash_validation(error)
            return render_template("pacientes/agendar_cita.html", **_quick_appointment_context()), 400
    return render_template("pacientes/agendar_cita.html", **_quick_appointment_context())


@pacientes.route("/<int:id>/registrar_proxima_cita", methods=["POST"])
@login_required
def registrar_proxima_cita(id):
    patient = db.session.get(Paciente, id)
    if not patient:
        flash("Paciente no encontrado.", "error")
        return redirect(url_for("pacientes.lista_pacientes_activos"))
    try:
        data = _appointment_values(request.form)
        with _APPOINTMENT_WRITE_LOCK:
            appointment = Cita.obtener_cita_pendiente(id)
            if not Cita.es_horario_disponible(
                data["fecha"], data["hora"], appointment.id if appointment else None
            ):
                raise ValidationError("El horario seleccionado ya no está disponible.")
            action = "cita.update" if appointment else "cita.create"
            if not appointment:
                appointment = Cita(paciente_id=id)
                db.session.add(appointment)
            appointment.fecha, appointment.hora, appointment.motivo = (
                data["fecha"],
                data["hora"],
                data["motivo"] or None,
            )
            appointment.estado, appointment.estatus, appointment.motivo_cancelacion = (
                "pendiente",
                "Programada",
                None,
            )
            db.session.flush()
            AuditLog.record(action, entity_type="cita", entity_id=appointment.id, metadata={"paciente_id": id})
            db.session.commit()
        flash("Cita guardada exitosamente.", "success")
    except ValidationError as error:
        _flash_validation(error)
    return redirect(url_for("pacientes.detalle_paciente", id=id))


@pacientes.route("/<int:id>/actualizar_cita/<int:cita_id>", methods=["POST"])
@login_required
def actualizar_cita(id, cita_id):
    appointment = db.session.get(Cita, cita_id)
    if not appointment or appointment.paciente_id != id:
        flash("Cita no encontrada.", "error")
        return redirect(url_for("pacientes.detalle_paciente", id=id))
    try:
        data = _appointment_values(request.form)
        with _APPOINTMENT_WRITE_LOCK:
            if not Cita.es_horario_disponible(data["fecha"], data["hora"], appointment.id):
                raise ValidationError("El horario seleccionado ya no está disponible.")
            appointment.fecha, appointment.hora, appointment.motivo = (
                data["fecha"],
                data["hora"],
                data["motivo"] or None,
            )
            AuditLog.record("cita.update", entity_type="cita", entity_id=appointment.id)
            db.session.commit()
        flash("Cita actualizada exitosamente.", "success")
    except ValidationError as error:
        _flash_validation(error)
    return redirect(url_for("pacientes.detalle_paciente", id=id))


@pacientes.route("/disponibilidad_horas", methods=["GET"])
@login_required
def disponibilidad_horas():
    try:
        target = date_value(request.args.get("fecha"), "Fecha", allow_future=True)
    except ValidationError as error:
        return jsonify({"success": False, "error": str(error)}), 400
    query = Cita.query.filter_by(fecha=target, estatus="Programada")
    exclude_appointment_id = request.args.get("excluir_cita_id", type=int)
    if exclude_appointment_id:
        query = query.filter(Cita.id != exclude_appointment_id)
    response = jsonify([item.hora.strftime("%H:%M") for item in query.order_by(Cita.hora).all()])
    response.headers["Cache-Control"] = "no-store"
    return response


@pacientes.route("/disponibilidad_citas", methods=["GET"])
@login_required
def disponibilidad_citas():
    try:
        target = date_value(request.args.get("fecha"), "Fecha", allow_future=True)
        raw_exclusion = request.args.get("excluir_cita_id")
        exclude_appointment_id = (
            integer(raw_exclusion, "Cita excluida", minimum=1) if raw_exclusion else None
        )
        if target < date.today():
            raise ValidationError("La fecha de cita no puede estar en el pasado.")
        if target > date.today() + timedelta(days=APPOINTMENT_MAX_FUTURE_DAYS):
            raise ValidationError("La disponibilidad sólo puede consultarse hasta dos años en el futuro.")
    except ValidationError as error:
        return jsonify({"success": False, "error": str(error)}), 400
    response = jsonify(
        {
            "success": True,
            "fecha": target.isoformat(),
            "horarios": Cita.obtener_disponibilidad_dia(
                target, excluir_cita_id=exclude_appointment_id
            ),
        }
    )
    response.headers["Cache-Control"] = "no-store"
    return response


@pacientes.route("/citas/<int:id>/cambiar-estatus", methods=["POST"])
@login_required
def cambiar_estatus_cita(id):
    appointment = db.session.get(Cita, id)
    if not appointment:
        return jsonify({"success": False, "error": "Cita no encontrada"}), 404
    attempted_status = ""
    try:
        payload = request.get_json(silent=True) or {}
        attempted_status = str(payload.get("estatus") or "")[:30]
        status = enum_value(attempted_status, "Estatus", ALLOWED_APPOINTMENT_STATUS)
        reason = clean_text(payload.get("motivo"), "Motivo", maximum=500)
        _validate_appointment_status_transition(appointment, status, reason)
        previous_status = appointment.estatus
        appointment.estatus = status
        appointment.estado = "completada"
        appointment.motivo_cancelacion = reason if status in {"Cancelada", "No Asistió"} else None
        AuditLog.record(
            "cita.status",
            entity_type="cita",
            entity_id=appointment.id,
            metadata={
                "status_anterior": previous_status,
                "status": status,
                "paciente_id": appointment.paciente_id,
                "con_observacion": bool(reason),
            },
        )
        db.session.commit()
        return jsonify({"success": True, "nuevo_estatus": status})
    except ValidationError as error:
        db.session.rollback()
        AuditLog.record(
            "cita.status",
            entity_type="cita",
            entity_id=appointment.id,
            outcome="denied",
            metadata={
                "status_actual": appointment.estatus,
                "status_solicitado": attempted_status,
                "paciente_id": appointment.paciente_id,
            },
        )
        db.session.commit()
        return jsonify({"success": False, "error": str(error)}), 400
