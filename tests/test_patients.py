import io
from datetime import date

from openpyxl import Workbook

from app import db_orm as db
from app.core.audit import AuditLog
from app.models.paciente import Paciente
from app.models.valoracion_antropometrica import ValoracionAntropometrica
from tests.conftest import csrf_from


def _patient_form(**overrides):
    data = {
        "nombre": "María Elena",
        "apellido_paterno": "García",
        "apellido_materno": "López",
        "genero": "mujer",
        "fecha_nacimiento": "1992-05-10",
        "telefono": "5512345678",
        "correo": "maria@example.test",
        "ciudad": "Ciudad de México",
    }
    data.update(overrides)
    return data


def _xlsx(*, invalid_second_row=False):
    workbook = Workbook()
    sheet = workbook.active
    sheet["M8"] = 1.70
    for row, visit in [(10, 1), (11, 2)] if invalid_second_row else [(10, 1)]:
        sheet.cell(row=row, column=12, value=visit)
        sheet.cell(row=row, column=13, value=date(2026, 1, visit))
        sheet.cell(row=row, column=14, value=70)
        sheet.cell(row=row, column=15, value="grasa 14")
        for column, value in {16: 80, 17: 95, 18: 32, 19: 95, 20: 55, 21: 36}.items():
            sheet.cell(row=row, column=column, value=value)
        sheet.cell(row=row, column=22, value="bc 5 tc 10 si 12 se 11 fem 15")
        sheet.cell(row=row, column=23, value="18%")
        sheet.cell(row=row, column=24, value="bad" if invalid_second_row and row == 11 else "120/80")
        sheet.cell(row=row, column=25, value=60)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def test_create_patient_validates_and_audits(app, client, login):
    login()
    token = csrf_from(client.get("/pacientes/nuevo"))
    response = client.post("/pacientes/nuevo", data={**_patient_form(), "csrf_token": token})
    assert response.status_code == 302
    with app.app_context():
        patient = Paciente.query.one()
        assert patient.telefono == "5512345678"
        assert AuditLog.query.filter_by(action="CREAR_PACIENTE", entity_id=patient.id).one()


def test_invalid_patient_is_not_persisted(app, client, login):
    login()
    token = csrf_from(client.get("/pacientes/nuevo"))
    response = client.post(
        "/pacientes/nuevo",
        data={**_patient_form(telefono="123"), "csrf_token": token},
        follow_redirects=True,
    )
    assert response.status_code == 200
    with app.app_context():
        assert Paciente.query.count() == 0


def test_patient_detail_collapses_uncaptured_optional_fields(app, client, login):
    login()
    with app.app_context():
        patient = Paciente(
            nombre="Sofía",
            apellido_paterno="Mendoza",
            genero="mujer",
            fecha_nacimiento=date(1988, 2, 15),
            telefono="5511112222",
            ciudad="Querétaro",
            status="activo",
        )
        db.session.add(patient)
        db.session.commit()
        patient_id = patient.id

    page = client.get(f"/pacientes/{patient_id}").get_data(as_text=True)

    assert 'data-patient-summary' in page
    assert 'data-patient-complementary-empty' in page
    assert 'data-patient-complementary-data' not in page
    assert "Información complementaria no capturada." in page
    assert "Completar datos" in page
    assert "Último pago" in page and "No registrado" in page
    assert "Siguiente cita" in page and "No programada" in page


def test_patient_detail_renders_only_captured_optional_fields(app, client, login):
    login()
    with app.app_context():
        patient = Paciente(
            nombre="Luis",
            apellido_paterno="Torres",
            genero="hombre",
            fecha_nacimiento=date(1979, 9, 20),
            telefono="5533334444",
            correo="luis@example.test",
            ciudad="León",
            ocupacion="Docente",
            direccion="Calle Salud 10",
            status="activo",
        )
        db.session.add(patient)
        db.session.commit()
        patient_id = patient.id

    page = client.get(f"/pacientes/{patient_id}").get_data(as_text=True)

    assert 'data-patient-complementary-data' in page
    assert 'data-patient-complementary-empty' not in page
    assert "luis@example.test" in page
    assert "Docente" in page
    assert "Calle Salud 10" in page
    assert ">Contacto de emergencia<" not in page
    assert ">Teléfono de emergencia<" not in page


def test_patient_search_ignores_case_accents_and_accepts_partial_name_terms(app, client, login):
    login()
    with app.app_context():
        db.session.add_all(
            [
                Paciente(
                    nombre="Sofía",
                    apellido_paterno="Núñez",
                    apellido_materno="Ocampo",
                    genero="mujer",
                    fecha_nacimiento=date(1991, 6, 12),
                    telefono="5511119999",
                    correo="sofia.nunez@example.test",
                    ciudad="Puebla",
                    status="activo",
                ),
                Paciente(
                    nombre="Mariana",
                    apellido_paterno="Torres",
                    genero="mujer",
                    fecha_nacimiento=date(1990, 2, 2),
                    telefono="5522229999",
                    correo="mariana@example.test",
                    ciudad="Puebla",
                    status="activo",
                ),
            ]
        )
        db.session.commit()

    for query in ("sofia", "SOFI", "sofi nune", "NUNEZ oca"):
        page = client.get("/pacientes/activos", query_string={"busqueda": query}).get_data(as_text=True)
        assert "Sofía" in page
        assert "Núñez Ocampo" in page
        assert "Mariana" not in page


def test_excel_import_is_atomic(app, client, login):
    login()
    with app.app_context():
        patient = Paciente.crear(**_patient_form(correo="maria@example.test"), status="activo")
        # The model expects a date, not the form representation.
        patient.fecha_nacimiento = date(1992, 5, 10)
        db.session.commit()
        patient_id = patient.id
    token = csrf_from(client.get(f"/pacientes/{patient_id}"))
    response = client.post(
        f"/pacientes/{patient_id}/cargar-excel",
        data={"csrf_token": token, "file": (_xlsx(invalid_second_row=True), "datos.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 400
    with app.app_context():
        assert ValoracionAntropometrica.query.count() == 0


def test_excel_import_accepts_valid_workbook(app, client, login):
    login()
    with app.app_context():
        patient = Paciente(
            nombre="María",
            apellido_paterno="García",
            apellido_materno="López",
            genero="mujer",
            fecha_nacimiento=date(1992, 5, 10),
            telefono="5512345678",
            correo="maria@example.test",
            ciudad="México",
            status="activo",
        )
        db.session.add(patient)
        db.session.flush()
        patient_id = patient.id
        previous_patient = Paciente(
            nombre="Paciente",
            apellido_paterno="Anterior",
            genero="otro",
            fecha_nacimiento=date(1985, 1, 1),
            telefono="5598765432",
            ciudad="México",
            status="activo",
        )
        db.session.add(previous_patient)
        db.session.flush()
        ValoracionAntropometrica.crear(
            previous_patient.id,
            {
                "numero_cita": 1,
                "fecha": date(2026, 1, 1),
                "motivo_consulta": "Atención previa del día",
            },
        )
        db.session.commit()
    token = csrf_from(client.get(f"/pacientes/{patient_id}"))
    response = client.post(
        f"/pacientes/{patient_id}/cargar-excel",
        data={"csrf_token": token, "file": (_xlsx(), "datos.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert response.get_json()["registros_procesados"] == 1
    with app.app_context():
        assert ValoracionAntropometrica.query.count() == 2
        imported = ValoracionAntropometrica.query.filter_by(paciente_id=patient_id).one()
        assert imported.numero_cita == 2
